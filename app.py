from flask import Flask, render_template, request, send_from_directory, jsonify
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import json
import tempfile
from datetime import datetime
import fitz  # PyMuPDF
import re
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Configuração para produção
if os.environ.get('VERCEL_ENV') == 'production':
    app.config['SERVER_NAME'] = os.environ.get('VERCEL_URL')

# Configuração para a pasta de saída
OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'pdf'}

# Modelo base de itens por sala
MODELO_BASE = {
    "Carteira": 50,
    "Mesa do Professor": 1,
    "Cadeira do Professor": 1,
    "Ventilador": 2,
    "Quadro Branco": 1,
    "Projetor": 1,
    "Computador": 1,
    "Armário": 1,
    "Lixeira": 1,
    "Ar Condicionado": 1,
    "Cortina": 2,
    "Luminária": 4
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extrair_tombamento_para_excel(pdf_path: str, excel_output: str):
    """
    Extrai dados de tombamento patrimonial de um PDF padronizado da Faculdade
    e exporta para uma planilha Excel com colunas: Código e Descrição.
    """
    regex_tombamento = re.compile(r"FACULDADE IBRA FABRAS\s+(\d{3})\s+(.+)", re.IGNORECASE)
    
    doc = fitz.open(pdf_path)
    registros = []

    for pagina in doc:
        texto = pagina.get_text()
        matches = regex_tombamento.findall(texto)
        for codigo, descricao in matches:
            registros.append((codigo.strip(), descricao.strip()))

    # Garantir que códigos únicos e ordenados sejam mantidos
    vistos = set()
    filtrado = []
    for codigo, descricao in registros:
        if codigo not in vistos:
            filtrado.append((codigo, descricao))
            vistos.add(codigo)

    # Criar DataFrame ordenado
    df = pd.DataFrame(filtrado, columns=["Código", "Descrição"])
    df["Código"] = df["Código"].astype(int)
    df = df.sort_values(by="Código")

    # Exportar como Excel
    df.to_excel(excel_output, index=False)
    return len(filtrado)  # Retorna o número de itens processados

@app.route('/')
def index():
    return render_template('index.html', modelo_base=json.dumps(MODELO_BASE))

@app.route('/gerar_preview', methods=['POST'])
def gerar_preview():
    # Recebe os dados do formulário
    data = request.get_json()
    num_salas = int(data.get('num_salas', 1))
    num_inicial = int(data.get('num_inicial', 1000))
    itens_personalizados = data.get('itens', MODELO_BASE)
    
    # Gera os dados do tombamento
    dados_tombamento = gerar_tombamento(num_salas, num_inicial, itens_personalizados)
    
    return jsonify({
        'dados': dados_tombamento
    })

@app.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    # Recebe os dados do formulário
    data = request.get_json()
    num_salas = int(data.get('num_salas', 1))
    num_inicial = int(data.get('num_inicial', 1000))
    itens_personalizados = data.get('itens', MODELO_BASE)
    
    # Gera os dados do tombamento
    dados_tombamento = gerar_tombamento(num_salas, num_inicial, itens_personalizados)
    
    # Cria o arquivo Excel
    nome_arquivo = f"tombamento_fabras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho_arquivo = os.path.join(OUTPUT_FOLDER, nome_arquivo)
    
    # Exporta para Excel
    exportar_para_excel(dados_tombamento, caminho_arquivo)
    
    return jsonify({
        'arquivo': nome_arquivo,
        'caminho': f"/download/{nome_arquivo}"
    })

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)

@app.route('/converter_pdf', methods=['POST'])
def converter_pdf():
    # Verifica se o arquivo foi enviado
    if 'pdfFile' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
        
    arquivo = request.files['pdfFile']
    
    # Verifica se o arquivo tem nome e é um PDF
    if arquivo.filename == '' or not arquivo.filename.lower().endswith('.pdf'):
        return jsonify({'erro': 'Arquivo inválido. Por favor, envie um PDF.'}), 400
    
    # Cria um nome seguro para o arquivo
    nome_seguro = secure_filename(arquivo.filename)
    nome_base = os.path.splitext(nome_seguro)[0]
          # Salva o arquivo temporariamente
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_pdf:
        arquivo.save(temp_pdf.name)
        caminho_pdf = temp_pdf.name

    try:
        # Nome do arquivo Excel de saída
        nome_arquivo = f"{nome_base}_convertido_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_arquivo = os.path.join(OUTPUT_FOLDER, nome_arquivo)
        
        # Extrai os dados do PDF para Excel usando nossa função personalizada
        num_itens = extrair_tombamento_para_excel(caminho_pdf, caminho_arquivo)
        
        # Formata o Excel para melhor visualização
        formatar_excel_convertido(caminho_arquivo)
        
        # Remove o arquivo temporário
        os.unlink(caminho_pdf)
        
        return jsonify({
            'arquivo': nome_arquivo,
            'caminho': f"/download/{nome_arquivo}",
            'num_itens': num_itens,
            'status': 'success'
        })
        
    except Exception as e:
        # Remove o arquivo temporário em caso de erro
        if os.path.exists(caminho_pdf):
            os.unlink(caminho_pdf)
        response = {'status': 'error', 'mensagem': str(e)}
        return jsonify(response), 500

def gerar_tombamento(num_salas, num_inicial, itens_personalizados):
    """Gera os dados do tombamento para o número especificado de salas."""
    dados = []
    numero_atual = num_inicial
    
    # Calcula o total de itens por sala
    total_itens_por_sala = sum(itens_personalizados.values())
    
    for sala in range(1, num_salas + 1):
        nome_sala = f"Sala {sala:02d}"
        
        # Adiciona cada item da sala
        for item, quantidade in itens_personalizados.items():
            for i in range(quantidade):
                dados.append({
                    "Sala": nome_sala,
                    "Item": item,
                    "Número": numero_atual,
                    "Descrição": f"{item} - {i+1 if quantidade > 1 else ''}".strip()
                })
                numero_atual += 1
    
    return dados

def exportar_para_excel(dados, caminho_arquivo):
    """Exporta os dados do tombamento para um arquivo Excel formatado."""
    # Cria um DataFrame com os dados
    df = pd.DataFrame(dados)
    
    # Cria um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tombamento Patrimonial"
    
    # Adiciona cabeçalho
    ws.merge_cells('A1:D1')
    ws['A1'] = "FABRAS - TOMBAMENTO PATRIMONIAL"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    
    # Adiciona data
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='right')
    
    # Adiciona cabeçalhos das colunas
    headers = ["Sala", "Item", "Número", "Descrição"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Adiciona os dados
    for row_idx, item in enumerate(dados, start=4):
        ws.cell(row=row_idx, column=1).value = item["Sala"]
        ws.cell(row=row_idx, column=2).value = item["Item"]
        ws.cell(row=row_idx, column=3).value = item["Número"]
        ws.cell(row=row_idx, column=4).value = item["Descrição"]
        
        # Adiciona bordas e alinhamento
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='center')
            
            # Centraliza a coluna de números
            if col == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajusta a largura das colunas
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20
    
    # Salva o arquivo
    wb.save(caminho_arquivo)
    return caminho_arquivo

def formatar_excel_convertido(caminho_arquivo):
    """Formata o arquivo Excel convertido do PDF para melhor visualização."""
    # Abre o arquivo Excel existente
    wb = openpyxl.load_workbook(caminho_arquivo)
    
    # Formata cada planilha
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formata o cabeçalho (primeira linha)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Formata as células de dados
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Tenta converter números para formato numérico
                if isinstance(cell.value, str):
                    try:
                        # Tenta converter para número se for um número
                        if re.match(r'^\d+$', cell.value):
                            cell.value = int(cell.value)
                        elif re.match(r'^\d+[.,]\d+$', cell.value.replace(',', '.')):
                            cell.value = float(cell.value.replace(',', '.'))
                    except (ValueError, TypeError):
                        pass
        
        # Ajusta a largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva o arquivo
    wb.save(caminho_arquivo)

if __name__ == '__main__':
    app.run(debug=True)